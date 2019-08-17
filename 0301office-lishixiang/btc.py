#!/usr/bin/env Python
# -*- coding:utf-8 -*-
# btc.py
# author: lishixiang


from openpyxl import load_workbook


"""
2. 下载表格 比特币价格表: btc.xlsx, 完成功能:
    - 根据第一列的数据, 按年进行拆分, 放到新的工作表, 比如2015年的所有数据放到新的sheet表,并命名2015, 2016放到新的sheet, 命名2016;
    - 按年拆分后的数据, 在数据的最后一行, 添加平均价格;
    - 做一个excel操作的类, 添加相应方法完成需求, 并有完整测试过程和结果文件.
"""

class BtcExcel():
    """
    一个操作比特币价格表的excel文档的类, 主要实现工作表的分类, 并在最后一行添加平均价格"""
    def __init__(self, filename):
        "初始化加载需要操作的excel文档"
        self.wb = load_workbook(filename)
        self.sheetnamelist = self.wb.sheetnames
        self.ws = self.wb[self.wb.sheetnames[0]]

    def create_sheet(self):
        """判断截取的年份在不在工作表中, 没有就创建然后添加, 有就直接添加进去"""
        for row in range(2, self.ws.max_row + 1):
            date = self.ws.cell(row=row, column=1).value
            price = self.ws.cell(row=row, column=2).value
            year = self.ws.cell(row=row, column=1).value[:4]
            if year not in self.sheetnamelist:  
                new_ws = 'wh' + year
                new_ws = self.wb.create_sheet(year)
                new_ws['A1'].value = self.ws['A1'].value
                new_ws['B1'].value = self.ws['B1'].value
                new_ws.cell(row=2, column=1).value = date
                new_ws.cell(row=2, column=2).value = price
                new_ws.column_dimensions['A'].width = 23
                new_ws.column_dimensions['B'].width = 12
                self.sheetnamelist.append(year)
            else:
                ws = self.wb[year]
                ws.cell(row=ws.max_row+1, column=1).value = date
                ws.cell(row=ws.max_row, column=2).value = price

    def average_price(self):
        "计算每年的平均价格"
        for sheet in self.sheetnamelist:
            if sheet != self.ws.title:
                ws = self.wb[sheet]
                max_row = ws.max_row
                max_cell = 'B' + str(max_row)
                ws.cell(row=ws.max_row+1, column=1).value = '平均值'
                ws.cell(row=ws.max_row, column=2).value = '=average(B2:%s)' %max_cell

    def save_excel(self, newfile):
        self.wb.save(newfile)
        return 'OK'


def main():
    wb = BtcExcel('btc.xlsx')
    wb.create_sheet()
    wb.average_price()
    print(wb.save_excel('btc_new.xlsx'))

if __name__ == '__main__':
    main()
    